import os
import re
import tempfile
import numpy as np
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font
import streamlit as st

TEMPLATE_FILE = "ХХХХ.99999.000 - Пример.xlsx"

# Базовые словари и списки (аналогичные вашему коду)
MACHINES_DB = {
    "Вакуумный насос": 0.73,
    "Пятиосевой обрабатывающий центр с ЧПУ QUADRO мод. \"ONYX SM2030\"": 17.0,
    "Wattsan A1 1325": 3.5,
    "Лазерный станок Wattsan 1610 LT": 4.0,
    "Фрезерный станок Wattsan 1325 А1": 3.5,
    "3-х осевой центр": 8.0,
    "5-ти осевой центр": 15.0,
    "Picaso Desinger X Series 2": 0.7,
    "Picaso XL": 1.2,
    "Laser Cut 6090": 2.5,
    "CNC Router 1325": 4.5,
    "Инжекционно-литьевая машина T150": 15.0,
    "Другой станок": 0.0
}

WORKSHOPS_LIST = [
    "Участок производства изделий из пластмасс",
    "Участок механической обработки",
    "Участок сборочных линий БВС",
    "Участок композитного производства"
]

UNITS_LIST = ["кг.", "шт.", "метр", "м2", "п.метр", "л.", "компл."]

POSITIONS_LIST = [
    "Обработчик изделий из пластмасс 2 разряда /Участок производства изделий из пластмасс /",
    "Оператор оборудования трехмерной печати /Участок производства изделий из пластмасс /",
    "Машинист вакуум-формовочной машины 2 разряда /Участок производства изделий из пластмасс /",
    "Литейщик пластмасс 2 разряда /Участок производства изделий из пластмасс /",
    "Мастер /Участок производства изделий из пластмасс /",
    "Наладчик инжекционно-литьевой машины /Участок производства изделий из пластмасс /",
    "Начальник участка /Участок производства изделий из пластмасс /",
    "Инженер-технолог /Участок производства изделий из пластмасс /",
    "Слесарь-сборщик летательных аппаратов 3 разряда /Участок механической обработки /",
    "Слесарь-сборщик летательных аппаратов 4 разряда /Участок механической обработки /",
    "Оператор лазерных установок 3 разряда /Участок механической обработки /",
    "Оператор станков с программным управлением 3 разряда /Участок механической обработки /",
    "Оператор станков с программным управлением 4 разряда /Участок механической обработки /",
    "Начальник смены /Участок механической обработки /",
    "Мастер /Участок механической обработки /"
]

st.set_page_config(page_title="Расчет себестоимости", layout="centered")
st.title("📦 Расчет себестоимости по STP-модели")

# Функция расчета габаритов STEP
def get_stp_bounding_box(file_bytes):
    try:
        cartesian_points = {}
        vertex_cp_ids = set()
        cp_pattern = re.compile(r"#(\d+)\s*=\s*CARTESIAN_POINT\s*\([^,]*\s*,\s*\(\s*(-?\d+\.?\d*(?:[eE][-+]?\d+)?)\s*,\s*(-?\d+\.?\d*(?:[eE][-+]?\d+)?)\s*,\s*(-?\d+\.?\d*(?:[eE][-+]?\d+)?)\s*\)\s*\)", re.IGNORECASE)
        vp_pattern = re.compile(r"#\d+\s*=\s*VERTEX_POINT\s*\([^,]*\s*,\s*#(\d+)\s*\)", re.IGNORECASE)

        text_content = file_bytes.decode('utf-8', errors='ignore')
        for line in text_content.splitlines():
            cp_match = cp_pattern.search(line)
            if cp_match:
                cartesian_points[cp_match.group(1)] = (
                    float(cp_match.group(2)), float(cp_match.group(3)), float(cp_match.group(4))
                )
            vp_match = vp_pattern.search(line)
            if vp_match:
                vertex_cp_ids.add(vp_match.group(1))

        valid_coords = [cartesian_points[pid] for pid in vertex_cp_ids if pid in cartesian_points]
        if len(valid_coords) < 4:
            valid_coords = list(cartesian_points.values())

        if len(valid_coords) >= 4:
            pts = np.array(valid_coords)
            dims = np.max(pts, axis=0) - np.min(pts, axis=0)
            sorted_dims = sorted([round(float(dims[0]), 1), round(float(dims[1]), 1), round(float(dims[2]), 1)], reverse=True)
            return f"Габариты (ДхШхВ): {sorted_dims[0]} x {sorted_dims[1]} x {sorted_dims[2]} мм"
    except Exception as e:
        print(e)
    return "Габариты: не определены"

# 1. Загрузка STP файла
uploaded_stp = st.file_uploader("Загрузить 3D модель (.stp / .step)", type=["stp", "step"])
model_name = ""
bounding_box_str = ""

if uploaded_stp:
    model_name = os.path.splitext(uploaded_stp.name)[0]
    bounding_box_str = get_stp_bounding_box(uploaded_stp.getvalue())
    st.success(f"Загружено: {uploaded_stp.name} | {bounding_box_str}")

st.divider()

# 2. Начальные параметры
st.subheader("1. Начальные параметры")
prod_qty = st.number_input("Количество продукции, шт.", min_value=1, value=1)
selected_workshops = st.multiselect("Цех (производство)", WORKSHOPS_LIST, default=[WORKSHOPS_LIST[0]])

st.divider()

# 3. Оборудование
st.subheader("2. Оборудование и время работы")
num_machines = st.number_input("Количество единиц оборудования", min_value=1, max_value=10, value=1)
machines_data = []

for i in range(int(num_machines)):
    cols = st.columns([3, 1, 1, 1])
    with cols[0]:
        m_choice = st.selectbox(f"Станок {i+1}", list(MACHINES_DB.keys()), key=f"mac_{i}")
    with cols[1]:
        m_power = st.text_input(f"Мощность {i+1}", value=str(MACHINES_DB[m_choice]), key=f"pow_{i}")
    with cols[2]:
        m_hours = st.number_input(f"Часы {i+1}", min_value=0, value=0, key=f"m_h_{i}")
    with cols[3]:
        m_mins = st.number_input(f"Мин {i+1}", min_value=0, max_value=59, value=0, key=f"m_m_{i}")
    machines_data.append({"name": m_choice, "power": m_power, "hours": m_hours, "mins": m_mins})

st.divider()

# 4. Сырье и материалы
st.subheader("3. Сырье и материалы")
num_mats = st.number_input("Количество материалов", min_value=1, max_value=10, value=1)
materials_data = []

for i in range(int(num_mats)):
    cols = st.columns([3, 1, 2])
    with cols[0]:
        mat_name = st.text_input(f"Материал {i+1} (название)", value="ABS GF12", key=f"mat_n_{i}")
    with cols[1]:
        mat_unit = st.selectbox(f"Ед. {i+1}", UNITS_LIST, key=f"mat_u_{i}")
    with cols[2]:
        mat_val = st.text_input(f"Расход на 1 шт. {i+1}", value="1.0", key=f"mat_v_{i}")
    materials_data.append({"name": mat_name, "unit": mat_unit, "consumption": mat_val})

st.divider()

# 5. Трудозатраты
st.subheader("4. Трудозатраты")
num_workers = st.number_input("Количество позиций рабочих", min_value=1, max_value=10, value=1)
workers_data = []

for i in range(int(num_workers)):
    cols = st.columns([3, 1, 1, 1])
    with cols[0]:
        w_pos = st.selectbox(f"Должность {i+1}", POSITIONS_LIST, key=f"w_p_{i}")
    with cols[1]:
        w_count = st.number_input(f"Человек {i+1}", min_value=1, value=1, key=f"w_c_{i}")
    with cols[2]:
        w_hours = st.number_input(f"Часы {i+1}", min_value=0, value=0, key=f"w_h_{i}")
    with cols[3]:
        w_mins = st.number_input(f"Мин {i+1}", min_value=0, max_value=59, value=0, key=f"w_m_{i}")
    workers_data.append({"position": w_pos, "count": w_count, "hours": w_hours, "mins": w_mins})

st.divider()

# Генерация Excel
if st.button("Сформировать и скачать Excel-файл", type="primary"):
    if not model_name:
        st.error("Пожалуйста, загрузите 3D-модель (.stp) в самом начале страницы!")
    elif not os.path.exists(TEMPLATE_FILE):
        st.error(f"Файл-шаблон '{TEMPLATE_FILE} ' не найден на сервере!")
    else:
        try:
            wb = openpyxl.load_workbook(TEMPLATE_FILE)
            formula_ref = "='Справочник для Запроса'!$C:$C"
            dv_positions = DataValidation(type="list", formula1=formula_ref, allow_blank=True)

            ws = wb["Запрос Данных для расчета"]
            ws.add_data_validation(dv_positions)

            no_fill = PatternFill(fill_type=None)
            regular_font = Font(name="Calibri", size=11, bold=False, color="000000")

            workshops_str = f'"{",".join([w.strip() for w in WORKSHOPS_LIST])}"'
            dv_workshop = DataValidation(type="list", formula1=workshops_str, allow_blank=True)
            ws.add_data_validation(dv_workshop)
            dv_workshop.add(ws['C5'])

            ws['C3'].value = model_name
            ws['C3'].font = regular_font
            ws['C3'].fill = no_fill

            ws['C4'].value = prod_qty
            ws['C4'].fill = no_fill
            
            ws['C5'].value = " / ".join(selected_workshops) if selected_workshops else ""
            ws['C5'].fill = no_fill

            mac_names = [m["name"] for m in machines_data if m["name"]]
            mac_powers = [str(m["power"]).replace(',', '.') for m in machines_data]
            mac_times = [f"{m['hours']} ч {m['mins']} м" for m in machines_data]

            ws['C6'].value = " / ".join(mac_names) if mac_names else ""
            ws['C6'].fill = no_fill
            ws['C7'].value = " / ".join(mac_powers) if mac_powers else "0"
            ws['C7'].fill = no_fill
            ws['C8'].value = " / ".join(mac_times) if mac_times else "0 ч 0 м"
            ws['C8'].fill = no_fill

            # Очистка блока материалов
            for r in range(10, 22):
                ws.cell(row=r, column=3).value = None
                ws.cell(row=r, column=3).fill = no_fill

            row_idx = 10
            for i, mat in enumerate(materials_data, 1):
                if row_idx > 21: break
                ws.cell(row=row_idx, column=1, value=f"5.{i}.")
                ws.cell(row=row_idx, column=2, value="Название сырья или материала")
                ws.cell(row=row_idx, column=3, value=mat["name"]).fill = no_fill

                ws.cell(row=row_idx+1, column=1, value=f"5.{i}.1.")
                ws.cell(row=row_idx+1, column=2, value="Ед.измерения материалов")
                ws.cell(row=row_idx+1, column=3, value=mat["unit"]).fill = no_fill

                ws.cell(row=row_idx+2, column=1, value=f"5.{i}.2.")
                ws.cell(row=row_idx+2, column=2, value="Расход сырья на 1 единицу изделия")
                ws.cell(row=row_idx+2, column=3, value=float(str(mat["consumption"]).replace(',', '.'))).fill = no_fill
                row_idx += 3

            # Очистка блока трудозатрат
            for r in range(23, 35):
                ws.cell(row=r, column=1).value = None
                ws.cell(row=r, column=2).value = None
                ws.cell(row=r, column=3).value = None
                ws.cell(row=r, column=3).fill = no_fill

            row_idx = 23
            for i, worker in enumerate(workers_data, 1):
                if row_idx > 34: break
                pos_cell = ws.cell(row=row_idx, column=3)
                dv_positions.add(pos_cell)

                ws.cell(row=row_idx, column=1, value=f"6.{i}.")
                ws.cell(row=row_idx, column=2, value="Должность производственного рабочего")
                pos_cell.value = worker["position"]
                pos_cell.fill = no_fill

                ws.cell(row=row_idx+1, column=1, value=f"6.{i}.1.")
                ws.cell(row=row_idx+1, column=2, value=f"Количество человек должности(указанной в пункте 6.{i}.) производят работу по выпуску (указанном в пункте 2.) изделий.")
                ws.cell(row=row_idx+1, column=3, value=worker["count"]).fill = no_fill

                ws.cell(row=row_idx+2, column=1, value=f"6.{i}.2.")
                ws.cell(row=row_idx+2, column=2, value=f"Количество часов на изготовление деталей всего выпуска одного вида изделия.")
                ws.cell(row=row_idx+2, column=3, value=f"{worker['hours']} ч {worker['mins']} м").fill = no_fill

                ws.cell(row=row_idx+3, column=1, value=f"6.{i}.3.")
                ws.cell(row=row_idx+3, column=2, value="(Для мех.цеха, при производстве изделий комплектом на одной раскладке).")
                ws.cell(row=row_idx+3, column=3, value=None).fill = no_fill
                row_idx += 4

            # Сохранение во временный файл для скачивания
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                wb.save(tmp.name)
                tmp_path = tmp.name

            with open(tmp_path, "rb") as file_download:
                st.download_button(
                    label="📥 Скачать готовый файл Excel",
                    data=file_download,
                    file_name=f"{model_name}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            st.success("Файл успешно сформирован!")
        except Exception as e:
            st.error(f"Ошибка при формировании файла: {e}")